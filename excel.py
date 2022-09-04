# only for creating a new excel file



import openpyxl as xl
wb=xl.load_workbook('transactions.xlsx')   #if we use double quotes then it shows error
sheet=wb['Sheet1']              #if we use double quotes then it shows error
cell=sheet['a1']
# or cell=sheet.cell(1,1)
print(cell.value)    #prints the value of cell a1
print(sheet.max_row) #prints the no of rows
print(sheet.max_column)

for row in range(2,sheet.max_row+1): # since range fn olny prints to num less than given range.so +!
    cell=sheet.cell(row,3)
    print(cell.value)     

    corrected_price=cell.value*0.9
    #new column
    corrected_price_cell=sheet.cell(row,4) # creating a new column since row 3.new one is 4.else it will overlap
    corrected_price_cell.value=corrected_price

wb.save('transactions2.xlsx')    #workbook save


#for both new file and the sheet



import openpyxl as xl
from openpyxl.chart import BarChart,Reference #class

def process_workbook(filename):
    wb=xl.load_workbook(filename) 
    sheet=wb['Sheet1']            

    for row in range(2,sheet.max_row+1): 
        cell=sheet.cell(row,3)
        print(cell.value)     

        corrected_price=cell.value*0.9
    
        corrected_price_cell=sheet.cell(row,4) 
        corrected_price_cell.value=corrected_price

    values=Reference(sheet,
                    min_row=2,
                    max_row=sheet.max_row,
                    min_col=4,
                    max_col=4)

    chart=BarChart()        #from barchart class
    chart.add_data(values)
    sheet.add_chart(chart,'e2')

    wb.save(filename)    # overwrite same file
